import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:C1')
sheet.merge_cells('A1:A13')
wb.save('example_merge.xlsx')