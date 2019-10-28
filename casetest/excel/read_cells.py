import openpyxl
wb = openpyxl.load_workbook('example.xlsx',data_only=True)
sheet = wb['SalesTax']
print(sheet['A2'])
print(sheet['A2'].value)
print(sheet['B2'].value)
print(sheet['C2'].value)
cell=sheet['A1']
print(cell.value)
print('Cell '+cell.coordinate+' is '+cell.value)
print(sheet.cell(row=2,column=1).value)
for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=3).value)
print(sheet.max_row,sheet.max_column)