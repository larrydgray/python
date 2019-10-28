# found this and thought it might come in handy
from traceback import format_exc
from pathlib import Path
from openpyxl import load_workbook
from pycel.excelcompiler import ExcelCompiler
import logging


class MESSAGES:
    CANT_EVALUATE_CELL = ("Couldn't evaluate cell {address}."
                          " Try to load and save xlsx file.")


class XLSXReader:
    """
    Provides (almost) universal interface to read xlsx file cell values.

    For formulae, tries to get their precomputed values or, if none,
    to evaluate them.
    """

    # Interface.

    def __init__(self, path: Path):
        self.__path = path
        self.__book = load_workbook(self.__path, data_only=False)

    def get_cell_value(self, address: str, sheet: str = None):
        # If no sheet given, work with active one.
        if sheet is None:
            sheet = self.__book.active.title

        # If cell doesn't contain a formula, return cell value.
        if not self.__cell_contains_formula(address, sheet):
            return self.__get_as_is(address, sheet)

        # If cell contains formula:
        # If there's precomputed value of the cell, return it.
        precomputed_value = self.__get_precomputed(address, sheet)
        if precomputed_value is not None:
            return precomputed_value

        # If not, try to compute its value from the formula and return it.
        # If failed, report an error and return empty value.
        try:
            computed_value = self.__compute(address, sheet)
        except:
            logging.warning(MESSAGES.CANT_EVALUATE_CELL
                            .format(address=address))
            logging.debug(format_exc())
            return None
        return computed_value                

    # Private part.

    def __cell_contains_formula(self, address, sheet):
        cell = self.__book[sheet][address]
        return cell.data_type is cell.TYPE_FORMULA

    def __get_as_is(self, address, sheet):
        # Return cell value.
        return self.__book[sheet][address].value

    def __get_precomputed(self, address, sheet):
        # If the sheet is not loaded yet, load it.
        if not hasattr(self, '__book_with_precomputed_values'):
            self.__book_with_precomputed_values = load_workbook(
                self.__path, data_only=True)
        # Return precomputed value.
        return self.__book_with_precomputed_values[sheet][address].value

    def __compute(self, address, sheet):
        # If the computation engine is not created yet, create it.
        if not hasattr(self, '__formulae_calculator'):
            self.__formulae_calculator = ExcelCompiler(self.__path)
        # Compute cell value.
        computation_graph = self.__formulae_calculator.gen_graph(
            address, sheet=sheet)
        return computation_graph.evaluate(f"{sheet}!{address}")